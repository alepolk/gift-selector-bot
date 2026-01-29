# Создай файл import_fixed.py

import sqlite3
from openpyxl import load_workbook

wb = load_workbook('gifts_fixed.xlsx')
sheet = wb.active

conn = sqlite3.connect('gifts.db')
cursor = conn.cursor()

# Очищаем старые данные
cursor.execute('DELETE FROM gifts')

# Импортируем
for row in range(2, sheet.max_row + 1):
    data = []
    for col in range(1, 12):
        data.append(sheet.cell(row=row, column=col).value)
    
    cursor.execute('''
        INSERT INTO gifts (id, name, price, description, budget_tags, gender_tags, age_tags, relationship_tags, occasion_tags, value_tags, interest_tags)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', data)

conn.commit()
conn.close()
wb.close()

print(f"Импортировано {sheet.max_row - 1} подарков!")