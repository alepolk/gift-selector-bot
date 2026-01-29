import sqlite3
from openpyxl import load_workbook

DB_PATH = "gifts.db"
EXCEL_PATH = "gifts_tagged_ALL_1-325.xlsx"

def import_gifts():
    """Импортирует подарки из Excel в базу данных"""
    
    # Открываем Excel
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    
    # Подключаемся к базе
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Очищаем таблицу перед импортом
    cursor.execute("DELETE FROM gifts")
    
    # Импортируем строки (пропускаем заголовок)
    count = 0
    for row in range(2, sheet.max_row + 1):
        gift_data = (
            sheet.cell(row=row, column=1).value,  # id
            sheet.cell(row=row, column=2).value,  # name
            sheet.cell(row=row, column=3).value,  # price
            sheet.cell(row=row, column=4).value,  # description
            sheet.cell(row=row, column=5).value,  # budget_tags
            sheet.cell(row=row, column=6).value,  # gender_tags
            sheet.cell(row=row, column=7).value,  # age_tags
            sheet.cell(row=row, column=8).value,  # relationship_tags
            sheet.cell(row=row, column=9).value,  # occasion_tags
            sheet.cell(row=row, column=10).value, # value_tags
            sheet.cell(row=row, column=11).value, # interest_tags
        )
        
        cursor.execute('''
            INSERT INTO gifts (id, name, price, description, budget_tags, 
                             gender_tags, age_tags, relationship_tags,
                             occasion_tags, value_tags, interest_tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', gift_data)
        count += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"✅ Импортировано подарков: {count}")

if __name__ == "__main__":
    import_gifts()